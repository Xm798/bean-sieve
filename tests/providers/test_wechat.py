"""Tests for WeChat Pay statement provider."""

from datetime import date, time
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook

from bean_sieve.providers import get_provider
from bean_sieve.providers.payment.wechat import WechatProvider


@pytest.fixture
def wechat_xlsx_file(tmp_path):
    """Create a temporary WeChat XLSX file."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None

    # Header rows (17 rows of metadata)
    header_data = [
        ("微信支付账单明细",),
        ("微信昵称：[测试用户]",),
        ("起始时间：[2025-01-01 00:00:00] 终止时间：[2025-01-31 23:59:59]",),
        ("导出类型：[全部]",),
        ("导出时间：[2025-02-01 10:00:00]",),
        (None,),  # Empty row
        ("共3笔记录",),
        ("收入：1笔 100.00元",),
        ("支出：2笔 50.00元",),
        ("中性交易：0笔 0.00元",),
        ("注：",),
        ("1. 测试提示1",),
        ("2. 测试提示2",),
        ("3. 测试提示3",),
        (None,),  # Empty row
        ("----------------------微信支付账单明细列表--------------------",),
        # Column headers
        (
            "交易时间",
            "交易类型",
            "交易对方",
            "商品",
            "收/支",
            "金额(元)",
            "支付方式",
            "当前状态",
            "交易单号",
            "商户单号",
            "备注",
        ),
    ]

    # Data rows
    data_rows = [
        (
            "2025-01-15 14:30:00",
            "商户消费",
            "瑞幸咖啡",
            "生椰拿铁",
            "支出",
            "¥9.90",
            "招商银行(1234)",
            "支付成功",
            "TX001",
            "M001",
            "/",
        ),
        (
            "2025-01-15 12:00:00",
            "商户消费",
            "公司食堂",
            "午餐",
            "支出",
            "¥15.00",
            "零钱",
            "支付成功",
            "TX002",
            "M002",
            "/",
        ),
        (
            "2025-01-14 10:00:00",
            "转账",
            "张三",
            "转账备注:感谢",
            "收入",
            "¥100.00",
            "/",
            "已收钱",
            "TX003",
            "/",
            "/",
        ),
    ]

    # Write all data
    for row in header_data + data_rows:
        ws.append(row)

    file_path = tmp_path / "wechat_test.xlsx"
    wb.save(file_path)
    return file_path


class TestWechatProvider:
    """Tests for WechatProvider."""

    def test_provider_registration(self):
        """Test that WeChat provider is properly registered."""
        provider = get_provider("wechat")
        assert isinstance(provider, WechatProvider)
        assert provider.provider_id == "wechat"
        assert provider.provider_name == "微信支付"
        assert ".xlsx" in provider.supported_formats
        assert ".csv" in provider.supported_formats

    def test_can_handle(self):
        """Test file format detection."""
        assert WechatProvider.can_handle(Path("wechat_statement.xlsx"))
        assert WechatProvider.can_handle(Path("微信账单.csv"))
        assert not WechatProvider.can_handle(Path("wechat_statement.eml"))
        assert not WechatProvider.can_handle(Path("statement.xlsx"))  # no keyword

    def test_parse_xlsx_transactions(self, wechat_xlsx_file):
        """Test parsing transactions from XLSX file."""
        provider = WechatProvider()
        transactions = provider.parse(wechat_xlsx_file)

        assert len(transactions) == 3

        # Check expense transaction
        expense = transactions[0]
        assert expense.date == date(2025, 1, 15)
        assert expense.time == time(14, 30, 0)
        assert expense.amount == Decimal("9.90")  # expense is positive
        assert expense.currency == "CNY"
        assert expense.payee == "瑞幸咖啡"
        assert expense.description == "生椰拿铁"
        assert expense.order_id == "TX001"
        assert expense.provider == "wechat"
        assert expense.is_expense

        # Check income transaction
        income = transactions[2]
        assert income.date == date(2025, 1, 14)
        assert income.amount == Decimal("-100.00")  # income is negative
        assert income.payee == "张三"
        assert income.is_income

    def test_metadata_extraction(self, wechat_xlsx_file):
        """Test that metadata is properly extracted."""
        provider = WechatProvider()
        transactions = provider.parse(wechat_xlsx_file)

        txn = transactions[0]
        assert txn.metadata["tx_type"] == "商户消费"
        assert txn.metadata["method"] == "招商银行(1234)"
        assert txn.metadata["status"] == "支付成功"
        assert txn.metadata["order_type"] == "支出"

    def test_statement_period_extraction(self, wechat_xlsx_file):
        """Test that statement period is extracted from header and set on transactions."""
        provider = WechatProvider()
        transactions = provider.parse(wechat_xlsx_file)

        # All transactions should have the statement period set
        for txn in transactions:
            assert txn.statement_period is not None
            assert txn.statement_period == (date(2025, 1, 1), date(2025, 1, 31))

    def test_neutral_transactions_filtered(self, tmp_path):
        """Test that neutral transactions (/) are filtered."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None

        # Minimal headers
        for _ in range(16):
            ws.append((None,))

        # Column header
        ws.append(
            (
                "交易时间",
                "交易类型",
                "交易对方",
                "商品",
                "收/支",
                "金额(元)",
                "支付方式",
                "当前状态",
                "交易单号",
                "商户单号",
                "备注",
            )
        )

        # One expense, one neutral
        ws.append(
            (
                "2025-01-15 10:00:00",
                "商户消费",
                "商家",
                "商品",
                "支出",
                "¥50.00",
                "余额",
                "成功",
                "TX001",
                "M001",
                "/",
            )
        )
        ws.append(
            (
                "2025-01-15 11:00:00",
                "零钱充值",
                "银行卡",
                "充值",
                "/",
                "¥100.00",
                "银行卡",
                "成功",
                "TX002",
                "/",
                "/",
            )
        )

        file_path = tmp_path / "wechat_neutral.xlsx"
        wb.save(file_path)

        provider = WechatProvider()
        transactions = provider.parse(file_path)

        # Only the expense should be included
        assert len(transactions) == 1
        assert transactions[0].order_id == "TX001"


class TestWechatCommissionHandling:
    """Tests for WeChat commission extraction from remarks."""

    def test_commission_extracted(self, tmp_path):
        """Test that commission is extracted from remarks and deducted from amount."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None

        # Headers
        for _ in range(16):
            ws.append((None,))

        ws.append(
            (
                "交易时间",
                "交易类型",
                "交易对方",
                "商品",
                "收/支",
                "金额(元)",
                "支付方式",
                "当前状态",
                "交易单号",
                "商户单号",
                "备注",
            )
        )

        # Transaction with commission in remarks
        ws.append(
            (
                "2025-01-15 10:00:00",
                "二维码收款",
                "客户",
                "收款",
                "收入",
                "¥100.00",
                "零钱",
                "已收钱",
                "TX001",
                "/",
                "服务费¥0.10",
            )
        )

        file_path = tmp_path / "wechat_commission.xlsx"
        wb.save(file_path)

        provider = WechatProvider()
        transactions = provider.parse(file_path)

        assert len(transactions) == 1
        txn = transactions[0]
        # Amount should be 100.00 - 0.10 = 99.90, negated for income
        assert txn.amount == Decimal("-99.90")
        assert txn.metadata["commission"] == "0.10"


class TestWechatCashWithdraw:
    """Tests for WeChat cash withdraw handling."""

    def test_cash_withdraw_treated_as_income(self, tmp_path):
        """Test that cash withdraw is treated as income (money to wallet)."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None

        # Headers
        for _ in range(16):
            ws.append((None,))

        ws.append(
            (
                "交易时间",
                "交易类型",
                "交易对方",
                "商品",
                "收/支",
                "金额(元)",
                "支付方式",
                "当前状态",
                "交易单号",
                "商户单号",
                "备注",
            )
        )

        # Cash withdraw transaction (marked as "/" in 收/支)
        ws.append(
            (
                "2025-01-15 10:00:00",
                "零钱提现",
                "银行卡",
                "提现",
                "/",
                "¥100.00",
                "零钱",
                "已转账",
                "TX001",
                "/",
                "/",
            )
        )

        file_path = tmp_path / "wechat_withdraw.xlsx"
        wb.save(file_path)

        provider = WechatProvider()
        transactions = provider.parse(file_path)

        assert len(transactions) == 1
        txn = transactions[0]
        # Cash withdraw should be expense (money leaving the account)
        assert txn.amount == Decimal("100.00")
        assert txn.is_expense

    def test_cash_withdraw_method_override(self, tmp_path):
        """Test that 零钱提现 overrides method from bank card to 零钱."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None

        for _ in range(16):
            ws.append((None,))
        ws.append(
            (
                "交易时间",
                "交易类型",
                "交易对方",
                "商品",
                "收/支",
                "金额(元)",
                "支付方式",
                "当前状态",
                "交易单号",
                "商户单号",
                "备注",
            ),
        )

        ws.append(
            (
                "2025-06-15 10:30:00",
                "零钱提现",
                "测试银行(0001)",
                "提现",
                "/",
                "¥200.00",
                "测试银行(0001)",
                "已转账",
                "TX001",
                "/",
                "/",
            ),
        )

        file_path = tmp_path / "wechat_withdraw_bank.xlsx"
        wb.save(file_path)

        provider = WechatProvider()
        transactions = provider.parse(file_path)

        assert len(transactions) == 1
        txn = transactions[0]
        assert txn.metadata["method"] == "零钱"
        assert txn.metadata["_withdrawal_target"] == "测试银行(0001)"
        assert txn.amount == Decimal("200.00")

    def test_merchant_withdraw_method_override(self, tmp_path):
        """Test that 经营账户提现 overrides method from bank card to 经营账户."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None

        for _ in range(16):
            ws.append((None,))
        ws.append(
            (
                "交易时间",
                "交易类型",
                "交易对方",
                "商品",
                "收/支",
                "金额(元)",
                "支付方式",
                "当前状态",
                "交易单号",
                "商户单号",
                "备注",
            ),
        )

        ws.append(
            (
                "2025-06-20 14:00:00",
                "经营账户提现",
                "示例银行(0002)",
                "提现",
                "/",
                "¥500.00",
                "示例银行(0002)",
                "已转账",
                "TX002",
                "/",
                "/",
            ),
        )

        file_path = tmp_path / "wechat_merchant_withdraw.xlsx"
        wb.save(file_path)

        provider = WechatProvider()
        transactions = provider.parse(file_path)

        assert len(transactions) == 1
        txn = transactions[0]
        assert txn.metadata["method"] == "经营账户"
        assert txn.metadata["_withdrawal_target"] == "示例银行(0002)"
        assert txn.amount == Decimal("500.00")


def test_card_last4_extracted_from_method_with_suffix():
    from bean_sieve.providers.payment.wechat import WechatProvider

    provider = WechatProvider()
    assert provider._extract_card_last4("招商银行信用卡(8355)") == "8355"
    assert provider._extract_card_last4("建设银行信用卡(0800)") == "0800"


def test_card_last4_none_for_wallet_methods():
    from bean_sieve.providers.payment.wechat import WechatProvider

    provider = WechatProvider()
    assert provider._extract_card_last4("零钱") is None
    assert provider._extract_card_last4("零钱通") is None
    assert provider._extract_card_last4("经营账户") is None
    assert provider._extract_card_last4("") is None
    assert provider._extract_card_last4(None) is None
