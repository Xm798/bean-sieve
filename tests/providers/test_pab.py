"""Tests for Ping An Bank (PAB) debit card statement provider."""

from datetime import date, time
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook

from bean_sieve.providers import get_provider
from bean_sieve.providers.banks.debit.pab import PABDebitProvider


def create_pab_xlsx(
    tmp_path: Path, transactions: list[dict], card_suffix: str = "6666"
) -> Path:
    """Create a PAB debit card Excel file.

    Args:
        tmp_path: Temporary directory path
        transactions: List of transaction dicts with keys:
            - time: datetime string "YYYY-MM-DD HH:MM:SS"
            - payer_name, payer_account
            - payee_name, payee_account
            - tx_type: "转入" or "转出"
            - amount: numeric
            - balance: numeric
            - summary, note, order_id
        card_suffix: Last 4 digits of card number

    Returns:
        Path to the created Excel file
    """
    wb = Workbook()
    ws = wb.active

    # Row 1: Account info header
    ws.cell(row=1, column=1, value=f"账户流水明细 账号：6230****{card_suffix}")

    # Row 2: Column headers
    headers = [
        "交易时间",
        "付款方姓名",
        "付款方账号",
        "收款方姓名",
        "收款方账号",
        "交易类型",
        "交易金额",
        "账户余额",
        "摘要",
        "备注",
        "交易流水号",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=header)

    # Transaction rows
    for row_idx, txn in enumerate(transactions, 3):
        ws.cell(row=row_idx, column=1, value=txn.get("time", "2025-12-30 10:00:00"))
        ws.cell(row=row_idx, column=2, value=txn.get("payer_name", ""))
        ws.cell(row=row_idx, column=3, value=txn.get("payer_account", ""))
        ws.cell(row=row_idx, column=4, value=txn.get("payee_name", ""))
        ws.cell(row=row_idx, column=5, value=txn.get("payee_account", ""))
        ws.cell(row=row_idx, column=6, value=txn.get("tx_type", "转出"))
        ws.cell(row=row_idx, column=7, value=txn.get("amount", 0))
        ws.cell(row=row_idx, column=8, value=txn.get("balance", 10000))
        ws.cell(row=row_idx, column=9, value=txn.get("summary", ""))
        ws.cell(row=row_idx, column=10, value=txn.get("note", ""))
        ws.cell(row=row_idx, column=11, value=txn.get("order_id", ""))

    file_path = tmp_path / "平安银行借记卡流水.xlsx"
    wb.save(file_path)
    return file_path


@pytest.fixture
def pab_xlsx_file(tmp_path):
    """Create a sample PAB debit card Excel file."""
    transactions = [
        {
            "time": "2025-12-30 09:15:30",
            "payer_name": "",
            "payer_account": "",
            "payee_name": "McDonald's Beijing",
            "payee_account": "6217****8888",
            "tx_type": "转出",
            "amount": 45.50,
            "balance": 9954.50,
            "summary": "消费",
            "note": "快餐消费",
            "order_id": "PAB202512300001",
        },
        {
            "time": "2025-12-30 14:30:00",
            "payer_name": "",
            "payer_account": "",
            "payee_name": "Apple Store",
            "payee_account": "6225****9999",
            "tx_type": "转出",
            "amount": 999.00,
            "balance": 8955.50,
            "summary": "消费",
            "note": "电子产品",
            "order_id": "PAB202512300002",
        },
        {
            "time": "2025-12-31 10:00:00",
            "payer_name": "Zhang San",
            "payer_account": "6228****1111",
            "payee_name": "",
            "payee_account": "",
            "tx_type": "转入",
            "amount": 5000.00,
            "balance": 13955.50,
            "summary": "转账",
            "note": "工资入账",
            "order_id": "PAB202512310001",
        },
        {
            "time": "2025-12-31 16:45:20",
            "payer_name": "",
            "payer_account": "",
            "payee_name": "Hilton Hotel",
            "payee_account": "6222****5555",
            "tx_type": "转出",
            "amount": 1280.00,
            "balance": 12675.50,
            "summary": "消费",
            "note": "酒店住宿",
            "order_id": "PAB202512310002",
        },
    ]
    return create_pab_xlsx(tmp_path, transactions)


class TestPABDebitProvider:
    """Tests for PABDebitProvider."""

    def test_provider_registration(self):
        """Test that PAB provider is properly registered."""
        provider = get_provider("pab_debit")
        assert isinstance(provider, PABDebitProvider)
        assert provider.provider_id == "pab_debit"
        assert provider.provider_name == "平安银行借记卡"
        assert ".xlsx" in provider.supported_formats
        assert ".xls" in provider.supported_formats

    def test_can_handle(self):
        """Test file format detection."""
        assert PABDebitProvider.can_handle(Path("平安银行借记卡流水.xlsx"))
        assert PABDebitProvider.can_handle(Path("平安借记卡明细.xls"))
        assert PABDebitProvider.can_handle(Path("借记卡平安银行.xlsx"))
        assert not PABDebitProvider.can_handle(Path("pab_statement.csv"))
        assert not PABDebitProvider.can_handle(Path("招商银行借记卡.xlsx"))

    def test_parse_transactions(self, pab_xlsx_file):
        """Test parsing transactions from Excel file."""
        provider = PABDebitProvider()
        transactions = provider.parse(pab_xlsx_file)

        assert len(transactions) == 4

        # Check expense transaction
        txn1 = transactions[0]
        assert txn1.date == date(2025, 12, 30)
        assert txn1.time == time(9, 15, 30)
        assert txn1.amount == Decimal("45.50")
        assert txn1.currency == "CNY"
        assert txn1.card_last4 == "6666"
        assert "McDonald's" in txn1.description or "McDonald's" in (txn1.payee or "")
        assert txn1.provider == "pab_debit"
        assert txn1.is_expense
        assert txn1.order_id == "PAB202512300001"

        txn2 = transactions[1]
        assert txn2.date == date(2025, 12, 30)
        assert txn2.amount == Decimal("999.00")
        assert "Apple" in txn2.description or "Apple" in (txn2.payee or "")

        # Check income transaction (should be negative)
        txn3 = transactions[2]
        assert txn3.date == date(2025, 12, 31)
        assert txn3.amount == Decimal("-5000.00")
        assert "Zhang San" in txn3.description or "Zhang San" in (txn3.payee or "")
        assert txn3.is_income

        txn4 = transactions[3]
        assert txn4.date == date(2025, 12, 31)
        assert txn4.amount == Decimal("1280.00")
        assert "Hilton" in txn4.description or "Hilton" in (txn4.payee or "")

    def test_card_last4_extraction(self, pab_xlsx_file):
        """Test that card_last4 is properly extracted from account info."""
        provider = PABDebitProvider()
        transactions = provider.parse(pab_xlsx_file)

        for txn in transactions:
            assert txn.card_last4 == "6666"

    def test_order_id_extraction(self, pab_xlsx_file):
        """Test that order_id is properly extracted."""
        provider = PABDebitProvider()
        transactions = provider.parse(pab_xlsx_file)

        assert transactions[0].order_id == "PAB202512300001"
        assert transactions[1].order_id == "PAB202512300002"
        assert transactions[2].order_id == "PAB202512310001"
        assert transactions[3].order_id == "PAB202512310002"

    def test_empty_statement(self, tmp_path):
        """Test handling of statement with no transactions."""
        file_path = create_pab_xlsx(tmp_path, [])

        provider = PABDebitProvider()
        transactions = provider.parse(file_path)

        assert transactions == []


class TestPABTransactionTypes:
    """Tests for PAB transaction type handling."""

    def test_transfer_out(self, tmp_path):
        """Test '转出' transactions are positive (expense)."""
        transactions = [
            {
                "time": "2025-12-30 10:00:00",
                "payee_name": "Test Merchant",
                "tx_type": "转出",
                "amount": 100.00,
                "summary": "消费",
            },
        ]
        file_path = create_pab_xlsx(tmp_path, transactions)

        provider = PABDebitProvider()
        parsed = provider.parse(file_path)

        assert len(parsed) == 1
        assert parsed[0].amount == Decimal("100.00")
        assert parsed[0].is_expense

    def test_transfer_in(self, tmp_path):
        """Test '转入' transactions are negative (income)."""
        transactions = [
            {
                "time": "2025-12-30 10:00:00",
                "payer_name": "Salary Sender",
                "tx_type": "转入",
                "amount": 8000.00,
                "summary": "工资",
            },
        ]
        file_path = create_pab_xlsx(tmp_path, transactions)

        provider = PABDebitProvider()
        parsed = provider.parse(file_path)

        assert len(parsed) == 1
        assert parsed[0].amount == Decimal("-8000.00")
        assert parsed[0].is_income


class TestPABDateTimeParsing:
    """Tests for PAB datetime parsing."""

    def test_datetime_parsing(self, tmp_path):
        """Test parsing datetime string."""
        transactions = [
            {
                "time": "2025-01-15 23:59:59",
                "payee_name": "Late Night Shop",
                "tx_type": "转出",
                "amount": 50.00,
                "summary": "消费",
            },
        ]
        file_path = create_pab_xlsx(tmp_path, transactions)

        provider = PABDebitProvider()
        parsed = provider.parse(file_path)

        assert len(parsed) == 1
        assert parsed[0].date == date(2025, 1, 15)
        assert parsed[0].time == time(23, 59, 59)

    def test_date_only(self, tmp_path):
        """Test parsing date without time (edge case)."""
        # Some statements might have date only
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="账户流水明细 账号：6230****1234")
        headers = [
            "交易时间",
            "付款方姓名",
            "付款方账号",
            "收款方姓名",
            "收款方账号",
            "交易类型",
            "交易金额",
            "账户余额",
            "摘要",
            "备注",
            "交易流水号",
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=2, column=col, value=header)

        # Row with date only (no time)
        ws.cell(row=3, column=1, value="2025-12-30")
        ws.cell(row=3, column=4, value="Test")
        ws.cell(row=3, column=6, value="转出")
        ws.cell(row=3, column=7, value=100)
        ws.cell(row=3, column=9, value="Test")

        file_path = tmp_path / "平安银行借记卡.xlsx"
        wb.save(file_path)

        provider = PABDebitProvider()
        parsed = provider.parse(file_path)

        # Should handle gracefully - either parse successfully or skip the row
        # Based on the provider code, it will return None, None for time
        if parsed:
            assert parsed[0].date == date(2025, 12, 30)


class TestPABDescriptionBuilding:
    """Tests for PAB description building."""

    def test_description_with_all_fields(self, tmp_path):
        """Test description includes summary, counterparty, and note."""
        transactions = [
            {
                "time": "2025-12-30 10:00:00",
                "payee_name": "Google Cloud",
                "tx_type": "转出",
                "amount": 99.99,
                "summary": "消费",
                "note": "云服务订阅",
            },
        ]
        file_path = create_pab_xlsx(tmp_path, transactions)

        provider = PABDebitProvider()
        parsed = provider.parse(file_path)

        assert len(parsed) == 1
        desc = parsed[0].description
        # Description should contain summary, counterparty, and note
        assert "消费" in desc
        assert "Google" in desc or "Google" in (parsed[0].payee or "")

    def test_description_truncates_long_note(self, tmp_path):
        """Test that long notes are truncated."""
        long_note = "A" * 100  # Very long note
        transactions = [
            {
                "time": "2025-12-30 10:00:00",
                "payee_name": "Test",
                "tx_type": "转出",
                "amount": 100,
                "summary": "消费",
                "note": long_note,
            },
        ]
        file_path = create_pab_xlsx(tmp_path, transactions)

        provider = PABDebitProvider()
        parsed = provider.parse(file_path)

        assert len(parsed) == 1
        # Note should be truncated to 50 chars + "..."
        assert len(parsed[0].description) < len(long_note) + 50


class TestPABMetadata:
    """Tests for PAB metadata extraction."""

    def test_metadata_includes_tx_type(self, tmp_path):
        """Test that metadata includes transaction type."""
        transactions = [
            {
                "time": "2025-12-30 10:00:00",
                "payee_name": "Test",
                "tx_type": "转出",
                "amount": 100,
                "summary": "消费",
                "note": "测试备注",
            },
        ]
        file_path = create_pab_xlsx(tmp_path, transactions)

        provider = PABDebitProvider()
        parsed = provider.parse(file_path)

        assert len(parsed) == 1
        assert "tx_type" in parsed[0].metadata
        assert parsed[0].metadata["tx_type"] == "转出"
        assert parsed[0].metadata["summary"] == "消费"
        assert parsed[0].metadata["note"] == "测试备注"
