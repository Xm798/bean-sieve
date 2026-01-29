"""Tests for export functionality."""

from datetime import date, time
from decimal import Decimal
from pathlib import Path

import pytest

from bean_sieve.core.export import (
    ALL_FIELDS,
    CORE_FIELDS,
    METADATA_FIELDS,
    export_csv,
    export_xlsx,
    generate_export_filename,
)
from bean_sieve.core.types import Transaction


@pytest.fixture
def sample_transactions():
    """Sample transactions for testing."""
    return [
        Transaction(
            date=date(2025, 1, 15),
            time=time(14, 30, 0),
            amount=Decimal("99.00"),
            currency="CNY",
            description="Coffee at Luckin",
            payee="瑞幸咖啡",
            card_last4="1234",
            order_id="ALI202501150001",
            metadata={
                "category": "餐饮",
                "method": "花呗",
                "status": "交易成功",
            },
        ),
        Transaction(
            date=date(2025, 1, 16),
            amount=Decimal("-500.00"),
            currency="CNY",
            description="Salary",
            payee="Company Inc",
            metadata={
                "tx_type": "转账",
            },
        ),
    ]


def test_core_fields():
    """Test that core fields are defined correctly."""
    assert "date" in CORE_FIELDS
    assert "time" in CORE_FIELDS
    assert "amount" in CORE_FIELDS
    assert "currency" in CORE_FIELDS
    assert "description" in CORE_FIELDS
    assert "payee" in CORE_FIELDS
    assert "card_last4" in CORE_FIELDS
    assert "order_id" in CORE_FIELDS


def test_metadata_fields():
    """Test that metadata fields whitelist is defined."""
    assert "category" in METADATA_FIELDS
    assert "peer_account" in METADATA_FIELDS
    assert "method" in METADATA_FIELDS
    assert "status" in METADATA_FIELDS
    assert "merchant_id" in METADATA_FIELDS
    assert "tx_type" in METADATA_FIELDS
    assert "summary" in METADATA_FIELDS
    assert "posting_date" in METADATA_FIELDS
    # Excluded fields
    assert "original_date" not in METADATA_FIELDS
    assert "section" not in METADATA_FIELDS


def test_all_fields_order():
    """Test that ALL_FIELDS = CORE_FIELDS + METADATA_FIELDS."""
    assert ALL_FIELDS == CORE_FIELDS + METADATA_FIELDS


def test_generate_export_filename():
    """Test export filename generation."""
    input_path = Path("/data/alipay_2025-01.csv")
    assert generate_export_filename(input_path, "csv") == Path(
        "/data/alipay_2025-01_export.csv"
    )
    assert generate_export_filename(input_path, "xlsx") == Path(
        "/data/alipay_2025-01_export.xlsx"
    )


def test_export_csv(sample_transactions, tmp_path):
    """Test CSV export."""
    output_path = tmp_path / "test_export.csv"
    export_csv(sample_transactions, output_path)

    assert output_path.exists()
    content = output_path.read_text(encoding="utf-8")
    lines = content.strip().split("\n")

    # Check header
    header = lines[0]
    assert "date" in header
    assert "amount" in header
    assert "category" in header
    assert "method" in header

    # Check data rows
    assert len(lines) == 3  # header + 2 transactions
    assert "2025-01-15" in lines[1]
    assert "99.00" in lines[1]
    assert "瑞幸咖啡" in lines[1]
    assert "餐饮" in lines[1]
    assert "花呗" in lines[1]


def test_export_xlsx(sample_transactions, tmp_path):
    """Test XLSX export."""
    output_path = tmp_path / "test_export.xlsx"
    export_xlsx(sample_transactions, output_path)

    assert output_path.exists()

    # Verify content with openpyxl
    from openpyxl import load_workbook

    wb = load_workbook(output_path)
    ws = wb.active
    assert ws is not None

    # Check header row
    header = [cell.value for cell in ws[1]]
    assert "date" in header
    assert "amount" in header
    assert "category" in header

    # Check data
    assert ws.max_row == 3  # header + 2 transactions
    assert ws.cell(row=2, column=1).value == "2025-01-15"


def test_export_csv_empty_metadata(tmp_path):
    """Test CSV export with empty metadata fields."""
    transactions = [
        Transaction(
            date=date(2025, 1, 15),
            amount=Decimal("100.00"),
            currency="CNY",
            description="Test",
        )
    ]
    output_path = tmp_path / "test.csv"
    export_csv(transactions, output_path)

    content = output_path.read_text(encoding="utf-8")
    lines = content.strip().split("\n")

    # Should have all columns even if empty
    header_count = len(lines[0].split(","))
    data_count = len(lines[1].split(","))
    assert header_count == len(ALL_FIELDS)
    assert data_count == len(ALL_FIELDS)
