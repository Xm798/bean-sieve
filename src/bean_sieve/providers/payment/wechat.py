"""WeChat Pay statement provider."""

from __future__ import annotations

import csv
import re
from datetime import date, datetime
from decimal import Decimal
from enum import StrEnum
from pathlib import Path

from openpyxl import load_workbook

from ...core.preset_rules import PresetRule, PresetRuleAction, PresetRuleCondition
from ...core.types import Transaction
from .. import register_provider
from ..base import BaseProvider


class WechatOrderType(StrEnum):
    """WeChat order type (收/支 column)."""

    EXPENSE = "支出"
    INCOME = "收入"
    NEUTRAL = "/"


class WechatTxType(StrEnum):
    """WeChat transaction type (交易类型 column)."""

    CONSUME = "商户消费"
    LUCKY_MONEY = "微信红包"
    TRANSFER = "转账"
    QR_INCOME = "二维码收款"
    QR_SEND = "扫二维码付款"
    GROUP_COLLECT = "群收款"
    REFUND = "退款"
    CASH_TO_CASH = "转入零钱通-来自零钱"
    INTO_CASH = "转入零钱通"
    CASH_IN = "零钱充值"
    CASH_WITHDRAW = "零钱提现"
    MERCHANT_WITHDRAW = "经营账户提现"
    CREDIT_CARD_REFUND = "信用卡还款"
    BUY_LICAITONG = "购买理财通"
    CASH_TO_LOOSE_CHANGE = "零钱通转出-到零钱"
    CASH_TO_OTHERS = "零钱通转出"
    FAMILY_CARD = "亲属卡交易"
    SPONSOR_CODE = "赞赏码"
    OTHER = "其他"
    DONATION = "分分捐"


# Number of header lines to skip in WeChat CSV/XLSX
WECHAT_HEADER_LINES = 17

# Regex to extract commission from remarks
COMMISSION_REGEX = re.compile(r"\d+\.\d{2}")

# Regex to extract 4-digit card suffix from method (e.g. "招商银行信用卡(8355)")
CARD_LAST4_REGEX = re.compile(r"\((\d{4})\)$")

# Regex to extract rebate from remarks (已优惠¥10.00)
REBATE_REGEX = re.compile(r"已优惠¥?(\d+\.?\d*)")

# Regex to extract statement period from header
# Format: 起始时间：[2025-01-01 00:00:00] 终止时间：[2025-01-31 23:59:59]
STATEMENT_PERIOD_REGEX = re.compile(
    r"起始时间[：:]\s*\[?(\d{4}-\d{2}-\d{2})\s+\d{2}:\d{2}:\d{2}\]?\s*"
    r"终止时间[：:]\s*\[?(\d{4}-\d{2}-\d{2})\s+\d{2}:\d{2}:\d{2}\]?"
)


@register_provider
class WechatProvider(BaseProvider):
    """
    Provider for WeChat Pay (微信支付) statement files.

    WeChat exports statements as either CSV or XLSX files with:
    - 17 header lines (metadata, statistics, notes)
    - Data columns:
      0: 交易时间 (Transaction time)
      1: 交易类型 (Transaction type)
      2: 交易对方 (Counterparty)
      3: 商品 (Item)
      4: 收/支 (Income/Expense)
      5: 金额(元) (Amount with ¥ prefix)
      6: 支付方式 (Payment method)
      7: 当前状态 (Status)
      8: 交易单号 (Transaction ID)
      9: 商户单号 (Merchant order ID)
      10: 备注 (Remarks)
    """

    provider_id = "wechat"
    provider_name = "微信支付"
    supported_formats = [".csv", ".xlsx"]
    filename_keywords = ["微信", "wechat", "weixin"]
    content_keywords = ["微信支付账单明细", "微信支付账单"]

    @staticmethod
    def _extract_card_last4(method: str | None) -> str | None:
        """Extract 4-digit card suffix from method string."""
        if not method:
            return None
        m = CARD_LAST4_REGEX.search(method)
        return m.group(1) if m else None

    def parse(self, file_path: Path) -> list[Transaction]:
        """Parse WeChat statement file (CSV or XLSX)."""
        suffix = file_path.suffix.lower()

        if suffix == ".xlsx":
            return self._parse_xlsx(file_path)
        else:
            return self._parse_csv(file_path)

    def _parse_xlsx(self, file_path: Path) -> list[Transaction]:
        """Parse WeChat XLSX statement file."""
        transactions = []

        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            raise ValueError(f"No active sheet in {file_path}")

        # Extract statement period from header rows
        statement_period: tuple[date, date] | None = None
        header_text = ""
        for line_num, row in enumerate(ws.iter_rows(values_only=True)):
            if line_num >= WECHAT_HEADER_LINES:
                break
            row_text = " ".join(str(cell) for cell in row if cell is not None)
            header_text += row_text + "\n"

        statement_period = self._extract_statement_period(header_text)

        for line_num, row in enumerate(ws.iter_rows(values_only=True)):
            # Skip header lines
            if line_num < WECHAT_HEADER_LINES:
                continue

            # Convert tuple to list and filter None values
            row_data = [str(cell) if cell is not None else "" for cell in row]

            # Skip empty rows
            if not row_data or all(not cell for cell in row_data):
                continue

            try:
                txn = self._parse_row(
                    row_data, file_path, line_num + 1, statement_period
                )
                if txn:
                    transactions.append(txn)
            except Exception as e:
                import logging

                logging.warning(
                    f"Failed to parse line {line_num + 1} in {file_path}: {e}"
                )
                continue

        wb.close()
        return transactions

    def _parse_csv(self, file_path: Path) -> list[Transaction]:
        """Parse WeChat CSV statement file."""
        transactions = []

        # Try UTF-8 first, then GBK
        for encoding in ["utf-8", "gbk"]:
            try:
                with open(file_path, encoding=encoding, newline="") as f:
                    content = f.read()
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ValueError(f"Cannot decode file {file_path}")

        # Remove tabs that WeChat adds to CSV
        content = content.replace("\t", "")

        # Extract statement period from header
        lines = content.splitlines()
        header_text = "\n".join(lines[:WECHAT_HEADER_LINES])
        statement_period = self._extract_statement_period(header_text)

        reader = csv.reader(lines)

        for line_num, row in enumerate(reader):
            # Skip header lines
            if line_num < WECHAT_HEADER_LINES:
                continue

            # Skip empty rows
            if not row or len(row) < 10:
                continue

            try:
                txn = self._parse_row(row, file_path, line_num + 1, statement_period)
                if txn:
                    transactions.append(txn)
            except Exception as e:
                import logging

                logging.warning(
                    f"Failed to parse line {line_num + 1} in {file_path}: {e}"
                )
                continue

        return transactions

    def _parse_row(
        self,
        row: list[str],
        file_path: Path,
        line_num: int,
        statement_period: tuple[date, date] | None = None,
    ) -> Transaction | None:
        """Parse a single row into a Transaction."""
        # Clean whitespace
        row = [field.strip() for field in row]

        # Extract fields
        tx_time_str = row[0]
        tx_type_str = row[1]
        peer = row[2]
        item = row[3]
        order_type_str = row[4]
        amount_str = row[5]
        method = row[6]
        status = row[7]
        order_id = row[8] if len(row) > 8 else ""
        merchant_id = row[9] if len(row) > 9 else ""
        remarks = row[10] if len(row) > 10 else ""

        # Parse order type
        order_type = self._get_order_type(order_type_str)
        withdrawal_target: str | None = None
        if order_type == WechatOrderType.NEUTRAL:
            # For withdrawals, treat as expense (money leaving the account)
            if tx_type_str in (
                WechatTxType.CASH_WITHDRAW.value,
                WechatTxType.MERCHANT_WITHDRAW.value,
            ):
                order_type = WechatOrderType.EXPENSE
                # method shows destination bank, but source is the wallet;
                # store destination for contra_account resolution, then
                # override method so txn.account resolves to the wallet asset
                source = (
                    "经营账户"
                    if tx_type_str == WechatTxType.MERCHANT_WITHDRAW.value
                    else "零钱"
                )
                if method and method != source:
                    withdrawal_target = method
                    method = source
            elif status in ("已存入零钱", "已存入经营账户"):
                order_type = WechatOrderType.INCOME
            else:
                return None

        # Parse datetime
        tx_datetime = datetime.strptime(tx_time_str, "%Y-%m-%d %H:%M:%S")

        # Parse amount (remove ¥ prefix)
        amount_str = amount_str.lstrip("¥").strip()
        amount = Decimal(amount_str)

        # Handle commission in remarks
        commission = Decimal("0")
        if "服务费" in remarks:
            match = COMMISSION_REGEX.search(remarks)
            if match:
                commission = Decimal(match.group())
                amount = amount - commission

        # Handle rebate in remarks (已优惠¥10.00)
        # ¥ symbol indicates CNY rebate
        rebate = Decimal("0")
        rebate_currency: str | None = None
        if "已优惠" in remarks and "¥" in remarks:
            match = REBATE_REGEX.search(remarks)
            if match:
                rebate = Decimal(match.group(1))
                rebate_currency = "CNY"

        # Determine sign based on order type
        # Convention: expense is positive, income is negative
        if order_type == WechatOrderType.INCOME:
            amount = -amount

        # Build description
        description = item if item and item != "/" else tx_type_str

        return Transaction(
            date=tx_datetime.date(),
            time=tx_datetime.time(),
            amount=amount,
            currency="CNY",
            description=description,
            payee=peer,
            order_id=order_id,
            card_last4=self._extract_card_last4(method),
            provider=self.provider_id,
            source_file=file_path,
            source_line=line_num,
            statement_period=statement_period,
            metadata={
                "tx_type": tx_type_str,
                "method": method,
                "status": status,
                "merchant_id": merchant_id,
                "remarks": remarks,
                "order_type": order_type_str,
                "commission": str(commission) if commission else None,
                "rebate": str(rebate) if rebate else None,
                "rebate_currency": rebate_currency,
                "_withdrawal_target": withdrawal_target,
            },
        )

    def _get_order_type(self, order_type_str: str) -> WechatOrderType:
        """Convert order type string to enum."""
        order_type_str = order_type_str.strip()
        try:
            return WechatOrderType(order_type_str)
        except ValueError:
            return WechatOrderType.NEUTRAL

    def _extract_statement_period(self, header_text: str) -> tuple[date, date] | None:
        """Extract statement period from header text.

        WeChat format: 起始时间：[2025-01-01 00:00:00] 终止时间：[2025-01-31 23:59:59]
        """
        match = STATEMENT_PERIOD_REGEX.search(header_text)
        if match:
            start_date = date.fromisoformat(match.group(1))
            end_date = date.fromisoformat(match.group(2))
            return (start_date, end_date)
        return None

    @classmethod
    def get_preset_rules(cls) -> list[PresetRule]:
        """Return preset rules for WeChat transactions."""
        return [
            # 已存入零钱
            PresetRule(
                rule_id="wechat_to_balance",
                name="零钱收款",
                provider="wechat",
                condition=PresetRuleCondition(
                    metadata={"status": r"已存入零钱"},
                ),
                action=PresetRuleAction(account_keyword="零钱"),
                priority=90,
            ),
            # 已存入经营账户
            PresetRule(
                rule_id="wechat_to_merchant",
                name="经营账户收款",
                provider="wechat",
                condition=PresetRuleCondition(
                    metadata={"status": r"已存入经营账户"},
                ),
                action=PresetRuleAction(account_keyword="经营账户"),
                priority=90,
            ),
            # 零钱提现
            PresetRule(
                rule_id="wechat_cash_withdraw",
                name="零钱提现",
                provider="wechat",
                condition=PresetRuleCondition(
                    metadata={"tx_type": r"^零钱提现$"},
                ),
                action=PresetRuleAction(
                    account_keyword="零钱",
                    contra_account_metadata_key="_withdrawal_target",
                ),
                priority=90,
            ),
            # 经营账户提现
            PresetRule(
                rule_id="wechat_merchant_withdraw",
                name="经营账户提现",
                provider="wechat",
                condition=PresetRuleCondition(
                    metadata={"tx_type": r"^经营账户提现$"},
                ),
                action=PresetRuleAction(
                    account_keyword="经营账户",
                    contra_account_metadata_key="_withdrawal_target",
                ),
                priority=90,
            ),
        ]
