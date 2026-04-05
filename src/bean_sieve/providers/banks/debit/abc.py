"""Agricultural Bank of China (农业银行) debit card statement provider."""

from __future__ import annotations

import logging
import re
import shutil
import tempfile
import warnings
from datetime import date, time
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from ....core.types import Transaction
from ... import register_provider
from ...base import BaseProvider

logger = logging.getLogger(__name__)


@register_provider
class ABCDebitProvider(BaseProvider):
    """
    Provider for Agricultural Bank of China (农业银行) debit card XLSX statements.

    Downloaded from: https://perbank.abchina.com/EbankSite/ebank/startup
    Path: 个人网上银行 → 账户明细查询 → 导出

    File format:
    - Format: XLSX (may have .xls extension)
    - Row 1: "账户明细查询"
    - Row 2: account info (账户, 户名, 起始/截止日期)
    - Row 3: column headers
    - Row 4+: data rows
    - Columns: 交易日期, 交易时间, 交易金额, 本次余额, 对方户名, 对方账号,
               交易行, 交易渠道, 交易类型, 交易用途, 交易摘要
    - All cells are string type
    """

    provider_id = "abc_debit"
    provider_name = "农业银行借记卡"
    supported_formats = [".xlsx", ".xls"]
    filename_pattern = re.compile(r"detail\d{8}")
    content_keywords = []  # Binary file, content detection not possible

    COL_TX_DATE = 0  # 交易日期
    COL_TX_TIME = 1  # 交易时间
    COL_AMOUNT = 2  # 交易金额 (signed: -expense, +income)
    COL_COUNTERPARTY_NAME = 4  # 对方户名
    COL_PURPOSE = 9  # 交易用途
    COL_SUMMARY = 10  # 交易摘要

    HEADER_KEYWORDS = ("交易日期", "交易时间")
    TITLE_MARKER = "账户明细查询"

    def parse(self, file_path: Path) -> list[Transaction]:
        """Parse ABC debit card XLSX statement."""
        wb = self._load_workbook(file_path)
        sheet = wb.active
        if sheet is None:
            raise ValueError(f"No active sheet in {file_path}")

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        # Validate title
        if not rows[0][0] or self.TITLE_MARKER not in str(rows[0][0]):
            logger.warning(
                "Row 1 does not contain '%s' in %s, skipping",
                self.TITLE_MARKER,
                file_path,
            )
            return []

        card_last4 = self._extract_card_last4(rows)
        header_idx = self._find_header_row(rows)

        transactions = []
        for row_idx, row in enumerate(rows[header_idx + 1 :], header_idx + 2):
            txn = self._parse_row(row, row_idx, card_last4, file_path)
            if txn:
                transactions.append(txn)

        return transactions

    def _load_workbook(self, file_path: Path):
        """Load workbook, handling .xls files that are actually xlsx format."""
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message="Workbook contains no default style",
                category=UserWarning,
            )
            if file_path.suffix.lower() == ".xls":
                tmp_path = None
                try:
                    with tempfile.NamedTemporaryFile(
                        suffix=".xlsx", delete=False
                    ) as tmp:
                        tmp_path = tmp.name
                        shutil.copy(file_path, tmp_path)
                    return load_workbook(tmp_path)
                finally:
                    if tmp_path:
                        Path(tmp_path).unlink(missing_ok=True)
            return load_workbook(file_path)

    @staticmethod
    def _extract_card_last4(rows: list) -> str | None:
        """Extract card last 4 digits from row 2 account info."""
        if len(rows) < 2 or not rows[1][0]:
            return None
        info = str(rows[1][0])
        match = re.search(r"\d{4}\*+(\d{4})", info)
        return match.group(1) if match else None

    def _find_header_row(self, rows: list) -> int:
        """Find header row by scanning for column header keywords."""
        for idx in range(min(10, len(rows))):
            cell = str(rows[idx][0] or "").strip()
            if any(kw in cell for kw in self.HEADER_KEYWORDS):
                return idx
        return 2  # fallback: row 3 (0-indexed = 2)

    def _parse_row(
        self,
        row: tuple,
        row_idx: int,
        card_last4: str | None,
        file_path: Path,
    ) -> Transaction | None:
        """Parse a single data row."""
        date_str = str(row[self.COL_TX_DATE] or "").strip()
        if not date_str or not date_str[0].isdigit():
            return None

        tx_date = self._parse_date(date_str)
        if tx_date is None:
            return None

        tx_time = self._parse_time(row[self.COL_TX_TIME])

        amount = self._parse_amount(row[self.COL_AMOUNT])
        if amount is None:
            return None

        counterparty = str(row[self.COL_COUNTERPARTY_NAME] or "").strip()
        purpose = str(row[self.COL_PURPOSE] or "").strip()
        summary = str(row[self.COL_SUMMARY] or "").strip()
        description = self._build_description(purpose, summary)

        return Transaction(
            date=tx_date,
            time=tx_time,
            amount=amount,
            currency="CNY",
            description=description,
            payee=counterparty or None,
            card_last4=card_last4,
            provider=self.provider_id,
            source_file=file_path,
            source_line=row_idx,
            metadata={"summary": summary},
        )

    @staticmethod
    def _parse_date(value: str) -> date | None:
        """Parse date from 'YYYY-MM-DD' string."""
        try:
            parts = value.split("-")
            return date(int(parts[0]), int(parts[1]), int(parts[2]))
        except (ValueError, IndexError):
            return None

    @staticmethod
    def _parse_time(value) -> time | None:
        """Parse time from 'HH:MM:SS' string."""
        time_str = str(value or "").strip()
        if not time_str:
            return None
        try:
            parts = time_str.split(":")
            return time(int(parts[0]), int(parts[1]), int(parts[2]))
        except (ValueError, IndexError):
            return None

    def _parse_amount(self, value) -> Decimal | None:
        """Parse signed amount string. Negate: source -=expense → positive, source +=income → negative."""
        try:
            cleaned = str(value or "").replace(",", "").strip()
            if not cleaned:
                return None
            d = Decimal(cleaned)
            if d == 0:
                return None
            return -d  # negate: source uses opposite sign convention
        except (ValueError, InvalidOperation):
            return None

    @staticmethod
    def _build_description(purpose: str, summary: str) -> str:
        """Build description from purpose and summary fields."""
        parts = [p for p in [purpose, summary] if p]
        return " | ".join(parts) if parts else "Unknown"
