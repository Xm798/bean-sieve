"""Ping An Bank (平安银行) debit card statement provider."""

from __future__ import annotations

import re
import shutil
import tempfile
from datetime import date, time
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from ....core.types import Transaction
from ... import register_provider
from ...base import BaseProvider


@register_provider
class PABDebitProvider(BaseProvider):
    """
    Provider for Ping An Bank (平安银行) debit card Excel statements.

    Parses .xls/.xlsx files exported from PAB online banking.
    Note: Some .xls files are actually xlsx format internally.
    """

    provider_id = "pab_debit"
    provider_name = "平安银行借记卡"
    supported_formats = [".xls", ".xlsx"]
    filename_keywords = ["平安银行", "平安借记卡"]
    content_keywords = []  # Binary file, can't check content

    # Column indices (0-based, after header row)
    COL_TIME = 0  # 交易时间
    COL_PAYER_NAME = 1  # 付款方姓名
    COL_PAYER_ACCOUNT = 2  # 付款方账号
    COL_PAYEE_NAME = 3  # 收款方姓名
    COL_PAYEE_ACCOUNT = 4  # 收款方账号
    COL_TYPE = 5  # 交易类型 (转入/转出)
    COL_AMOUNT = 6  # 交易金额
    COL_BALANCE = 7  # 账户余额
    COL_SUMMARY = 8  # 摘要
    COL_NOTE = 9  # 备注
    COL_ORDER_ID = 10  # 交易流水号

    def parse(self, file_path: Path) -> list[Transaction]:
        """Parse PAB debit card Excel statement."""
        wb = self._load_workbook(file_path)
        sheet = wb.active

        transactions = []
        card_suffix = self._extract_card_suffix(sheet)

        # Skip header rows (row 1: account info, row 2: column headers)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), 3):
            txn = self._parse_row(row, row_idx, card_suffix, file_path)
            if txn:
                transactions.append(txn)

        return transactions

    def _load_workbook(self, file_path: Path):
        """Load workbook, handling .xls files that are actually xlsx format."""
        if file_path.suffix.lower() == ".xls":
            # Some .xls files are actually xlsx format internally
            # Copy to temp file with .xlsx extension for openpyxl
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                shutil.copy(file_path, tmp.name)
                return load_workbook(tmp.name)
        return load_workbook(file_path)

    def _extract_card_suffix(self, sheet) -> str | None:
        """Extract card suffix from the first row account info."""
        first_row = sheet.cell(1, 1).value
        if first_row:
            # Match pattern like "6230****6666"
            match = re.search(r"\d{4}\*+(\d{4})", first_row)
            if match:
                return match.group(1)
        return None

    def _parse_row(
        self,
        row: tuple,
        row_idx: int,
        card_suffix: str | None,
        file_path: Path,
    ) -> Transaction | None:
        """Parse a single transaction row."""
        if not row or not row[self.COL_TIME]:
            return None

        # Parse datetime
        time_str = str(row[self.COL_TIME])
        tx_date, tx_time = self._parse_datetime(time_str)
        if not tx_date:
            return None

        # Parse amount and determine sign
        amount_raw = row[self.COL_AMOUNT]
        if amount_raw is None:
            return None
        amount = Decimal(str(amount_raw))

        tx_type = row[self.COL_TYPE]
        # "转入" = income (negative), "转出" = expense (positive)
        if tx_type == "转入":
            amount = -amount

        # Build description
        summary = row[self.COL_SUMMARY] or ""
        note = row[self.COL_NOTE] or ""
        counterparty = self._get_counterparty(row, tx_type)
        description = self._build_description(summary, counterparty, note)

        # Get order_id
        order_id = str(row[self.COL_ORDER_ID]) if row[self.COL_ORDER_ID] else None

        return Transaction(
            date=tx_date,
            time=tx_time,
            amount=amount,
            currency="CNY",
            description=description,
            payee=counterparty,
            card_suffix=card_suffix,
            order_id=order_id,
            provider=self.provider_id,
            source_file=file_path,
            source_line=row_idx,
            metadata={
                "tx_type": tx_type,
                "summary": summary,
                "note": note,
            },
        )

    def _parse_datetime(self, time_str: str) -> tuple[date | None, time | None]:
        """Parse datetime string like '2025-12-30 17:10:07'."""
        try:
            parts = time_str.split(" ")
            date_parts = parts[0].split("-")
            tx_date = date(int(date_parts[0]), int(date_parts[1]), int(date_parts[2]))

            tx_time = None
            if len(parts) > 1:
                time_parts = parts[1].split(":")
                tx_time = time(
                    int(time_parts[0]), int(time_parts[1]), int(time_parts[2])
                )

            return tx_date, tx_time
        except (ValueError, IndexError):
            return None, None

    def _get_counterparty(self, row: tuple, tx_type: str) -> str:
        """Get counterparty name based on transaction type."""
        if tx_type == "转入":
            # For incoming, counterparty is payer
            return row[self.COL_PAYER_NAME] or ""
        else:
            # For outgoing, counterparty is payee
            return row[self.COL_PAYEE_NAME] or ""

    def _build_description(self, summary: str, counterparty: str, note: str) -> str:
        """Build transaction description from available fields."""
        parts = []
        if summary:
            parts.append(summary)
        if counterparty:
            parts.append(counterparty)
        if note and note != counterparty:
            # Truncate long notes
            short_note = note[:50] + "..." if len(note) > 50 else note
            parts.append(short_note)
        return " | ".join(parts) if parts else "Unknown"
