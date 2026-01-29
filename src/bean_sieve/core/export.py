"""Export transactions to CSV/XLSX format."""

from __future__ import annotations

import csv
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from .types import Transaction

# Core fields from Transaction model
CORE_FIELDS = [
    "date",
    "time",
    "amount",
    "currency",
    "description",
    "payee",
    "card_last4",
    "order_id",
]

# Metadata fields to include (whitelist)
METADATA_FIELDS = [
    "category",
    "peer_account",
    "method",
    "status",
    "merchant_id",
    "tx_type",
    "summary",
    "posting_date",
]

ALL_FIELDS = CORE_FIELDS + METADATA_FIELDS


def _transaction_to_row(txn: Transaction) -> dict[str, str]:
    """Convert a transaction to a row dict for export."""
    row: dict[str, str] = {
        "date": txn.date.isoformat(),
        "time": txn.time.isoformat() if txn.time else "",
        "amount": str(txn.amount),
        "currency": txn.currency,
        "description": txn.description,
        "payee": txn.payee or "",
        "card_last4": txn.card_last4 or "",
        "order_id": txn.order_id or "",
    }

    # Add metadata fields
    for field in METADATA_FIELDS:
        row[field] = str(txn.metadata.get(field, "")) if txn.metadata.get(field) else ""

    return row


def export_csv(transactions: list[Transaction], output_path: Path) -> None:
    """Export transactions to CSV format (UTF-8 without BOM)."""
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=ALL_FIELDS)
        writer.writeheader()
        for txn in transactions:
            writer.writerow(_transaction_to_row(txn))


def export_xlsx(transactions: list[Transaction], output_path: Path) -> None:
    """Export transactions to XLSX format."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Transactions"

    # Write header
    for col, field in enumerate(ALL_FIELDS, 1):
        ws.cell(row=1, column=col, value=field)

    # Write data
    for row_idx, txn in enumerate(transactions, 2):
        row_data = _transaction_to_row(txn)
        for col, field in enumerate(ALL_FIELDS, 1):
            ws.cell(row=row_idx, column=col, value=row_data[field])

    wb.save(output_path)


def generate_export_filename(input_path: Path, format: str) -> Path:
    """Generate output filename from input filename."""
    stem = input_path.stem
    return input_path.parent / f"{stem}_export.{format}"
